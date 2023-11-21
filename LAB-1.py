#CODE WITH FUNCTIONS THAT READS DATA FROM EXCEL SHEET AND CALCULATES THE REUIRED VARIENCE,CORRELATION AND REGRESSION USING RECURSION
#AND HIGHER ORDER FUNCTIONS AT APPROPRIATE PLACES. IT RETURNS VALUES THAT ARE WRITTEN BACK INTO EXCEL SHEET AGAIN
from functools import reduce
from math import sqrt
import openpyxl
wb = openpyxl.load_workbook("C:\\Users\\arusu\\Desktop\\VIT SEM-5 MATERIALS\\ADVANCED PYTHON PROGRAMMING\\LAB\\LAB-1\\LAB-1 EXCEL.xlsx")
sheet = wb.active

#USAGE OF REDUCE
def sum_highord(sum_list):
    return reduce(lambda m,n: m+n,sum_list)

#SWAP FUNCTION FOR BUBBLE SORT
def swap(A, i, j):
 
    temp = A[i]
    A[i] = A[j]
    A[j] = temp


#BUBBLE SORT WITH RECURSIVE CALL i.e., RECURSIVE FUNCTION
#RECURSION USED HERE
def bubbleSort(A, n):
 
    for i in range(n - 1):
        if A[i] > A[i + 1]:
            swap(A, i, i + 1)
 
    if n - 1 > 1:
        bubbleSort(A, n - 1)


#FUNCTION TO CALCULATE VARIENCE BETWEEN TWO ROWS
def variance(data):
    n = len(data)
    mean = sum_highord(data) / n
    deviation = [(x - mean) ** 2 for x in data]
    varince = sum_highord(deviation) / n
    return varince


#FUNCTION TO CALCULATE MEAN
def mean(mean_list):
    return sum_highord(mean_list) / len(mean_list)


#FUNCTION TO CALCULATE CORRELATION BETWEEN TWO ROWS
#LAMBDA FUNCTIONS USED INSIDE THIS
def correl_data(list1, list2):
    n = len(list1)
    '# finding the mean'
    x_mean = mean(list1)
    y_mean = mean(list2)
    # Mathematical operations in lambda functions
    p = lambda x, y: x - y
    c = lambda x, y: x * y    
    s = lambda x: x ** 2
    x_minus_mean = [p(x, x_mean) for x in list1]
    y_minus_mean = [p(y, y_mean) for y in list2]
    xy_multiply = [c(x_minus_mean[x], y_minus_mean[x]) for x in range(n)]
    x_square = [s(x_minus_mean[x]) for x in range(n)]
    y_square = [s(y_minus_mean[y]) for y in range(n)]
    denominator = sqrt(sum_highord(x_square) * sum_highord(y_square))
    numerator = sum_highord(xy_multiply)
    if numerator == 0 or denominator == 0:
        return 0
    return numerator / denominator


def num_select(rl):
    cnt = 0
    for j in range(len(rl)):
        if type(rl[j]) != str:
            cnt += 1
    return cnt


def call_select(row_list):
    count = num_select(row_list)
    r = [1] * count
    k = 0
    for j in range(len(row_list)):
        if type(row_list[j]) != str:
            r[k] = row_list[j]
            k += 1
    return r


#FUNCTION TO CALCULATE REGRESSION COORDINATES Y FOR A GIVEN X VALUES
#IF THE DENOMINATOR OR NUMERATOR OF A COEFFICIENT BECOMES 0, THW WHOLE COEFFICIENT BECOMES 0 
#HIGHER ORDER FUNCTIONS IS USED INSIDE THESE FUNCTIONS
def divider(x):
    def sub_divider(y):
        return x / y
    return sub_divider

def simple_regression (x, y):
    x_into_y = []
    x_into_x = []
    n = len(x)
    for i in range(len(x)):
        mul_val = x[i] * y[i]
        x_into_y.append(mul_val)
    for i in range(len(x)):
        x_into_x.append(x[i] * x[i])
    sum_x_into_y = sum_highord(x_into_y)
    sum_xlist = sum_highord(x)
    sum_ylist = sum_highord(y)
    sum_x_into_x = sum_highord(x_into_x)
    t = divider((n * sum_x_into_y) - (sum_xlist * sum_ylist))
    if ((n * sum_x_into_x) - (sum_xlist * sum_xlist)) == 0 or ((n * sum_x_into_y) - (sum_xlist * sum_ylist))==0 :
        b=0
    else:    
        b = ((n * sum_x_into_y) - (sum_xlist * sum_ylist)) / ((n * sum_x_into_x) - (sum_xlist * sum_xlist))
    if ((sum_ylist) - (b * sum_xlist)) == 0:
        a=0
    else:    
        a = ((sum_ylist) - (b * sum_xlist)) / (n)
    y_val = []
    for i in x:
        q = (b * i) + a
        y_val.append(float("{:.2f}".format(q)))
    return y_val


max_column = sheet.max_column
max_row = sheet.max_row
var = []


for i in range(2, max_row+1):
    row = sheet[i]
    row_list = [row[x].value for x in range(max_column)]
    r = call_select(row_list)
    bubbleSort(r,len(r))
    var_list = variance(r)
    var.append(var_list)


def append_cellvalues(c, val, l):
    sheet.cell(row=1, column=c + 2, value=val)
    for index, value in enumerate(l):
        sheet.cell(row=index + 2, column=c + 2, value=value)
    return c+2


col_val = append_cellvalues(max_column, "Variance", var)
for r in range(2, max_row):
    row_1 = sheet[r]
    row_list_1 = [row_1[x].value for x in range(max_column)]
    cor = []
    for r2 in range(r+1, max_row+1):
        row2 = sheet[r2]
        row_list_2 = [row2[x].value for x in range(max_column)]
        r_1 = call_select(row_list_1)
        r_2 = call_select(row_list_2)
        cor_val = correl_data(r_1, r_2)
        cor.append(cor_val)
    iop = append_cellvalues(col_val, "Co-Relation R-"+str(r), cor)
    col_val += 1


for loop_var in range(2, max_row):
    row_1 = sheet[loop_var]
    row_list_1 = [row_1[x].value for x in range(max_column)]
    reg_lst = []
    for r2 in range(loop_var+1, max_row+1):
        row2 = sheet[r2]
        row_list_2 = [row2[x].value for x in range(max_column)]
        r_1 = call_select(row_list_1)
        r_2 = call_select(row_list_2)
        regression = simple_regression(r_1, r_2)
        for f in range(len(regression)):
            reg_lst.append(regression[f])
        reg_app = append_cellvalues(col_val, "Regression R-"+str(loop_var), reg_lst)
        col_val += 1
        reg_lst.clear()


wb.save("C:\\Users\\arusu\\Desktop\\VIT SEM-5 MATERIALS\\ADVANCED PYTHON PROGRAMMING\\LAB\\LAB-1\\LAB-1 EXCEL.xlsx")