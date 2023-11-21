import openpyxl
wb = openpyxl.load_workbook("C:\\Users\\arusu\\Desktop\\VIT SEM-7 MATERIALS\\BLOCKCHAIN TECHNOLOGIES\\J Component\\CODE\\dataset.xlsx")
sheet = wb.active

def append_cellvalues(maxrow,value,colv):
    for index,value in enumerate(value):
        sheet.cell(row=index +maxrow, column=colv,value=value)

def appending(private_key_list,public_key_list,address_list,private_key_len,public_key_len,address_len):
    max_row = sheet.max_row+1
    append_cellvalues(max_row,private_key_list,1)
    append_cellvalues(max_row,public_key_list,2)
    append_cellvalues(max_row,address_list,3)
    append_cellvalues(max_row,private_key_len,4)
    append_cellvalues(max_row,public_key_len,5)
    append_cellvalues(max_row,address_len,6)
    wb.save("C:\\Users\\arusu\\Desktop\\VIT SEM-7 MATERIALS\\BLOCKCHAIN TECHNOLOGIES\\J Component\\CODE\\dataset.xlsx")